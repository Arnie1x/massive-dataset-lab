import pandas as pd
import openpyxl as xlsx

jsonObj = pd.read_json(path_or_buf='dataset/data/en-US.jsonl', lines=True).get(['id', 'utt', 'annot_utt'])
jsonObjSw = pd.read_json(path_or_buf='dataset/data/sw-KE.jsonl', lines=True).get(['id', 'utt', 'annot_utt'])

# print(jsonObj.loc[jsonObj['id'] == 1])
# print(jsonObjSw.loc[jsonObj['id'] == 1])
# print(jsonObj)

workbook = xlsx.Workbook()
worksheet = workbook.active
# Adding Headers to the Sheet
worksheet['A1'] = 'id'
worksheet['B1'] = 'utt'
worksheet['C1'] = 'annot_utt'
worksheet['D1'] = 'sw_utt'
worksheet['E1'] = 'sw_annot_utt'

for index, row in jsonObj.iterrows():
    # Column 1
    var = worksheet.cell(row=index + 2, column=1)
    var.value = row['id'] + 1
    # Column 2
    var = worksheet.cell(row=index + 2, column=2)
    var.value = row['utt']
    # Column 3
    var = worksheet.cell(row=index + 2, column=3)
    var.value = row['annot_utt']

for index, row in jsonObjSw.iterrows():
    # Column 2
    var = worksheet.cell(row=index + 2, column=4)
    var.value = row['utt']
    # Column 3
    var = worksheet.cell(row=index + 2, column=5)
    var.value = row['annot_utt']

workbook.save('outputs/en-sw.xlsx')
