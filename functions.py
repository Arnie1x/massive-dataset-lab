import pandas as pd
import openpyxl as xlsx


def english_translations(file_path, language_abbv):
    df_pivot = pd.read_json(path_or_buf='dataset/data/en-US.jsonl', lines=True).get(['id', 'utt', 'annot_utt'])
    df_lang = pd.read_json(path_or_buf=file_path, lines=True).get(['id', 'utt', 'annot_utt'])

    workbook = xlsx.Workbook()
    worksheet = workbook.active
    # Adding Headers to the Sheet
    worksheet['A1'] = 'id'
    worksheet['B1'] = 'utt'
    worksheet['C1'] = 'annot_utt'
    worksheet['D1'] = 'sw_utt'
    worksheet['E1'] = 'sw_annot_utt'

    # Adding english data to the sheet
    for index, row in df_pivot.iterrows():
        # Column 1
        var = worksheet.cell(row=index + 2, column=1)
        var.value = row['id'] + 1
        # Column 2
        var = worksheet.cell(row=index + 2, column=2)
        var.value = row['utt']
        # Column 3
        var = worksheet.cell(row=index + 2, column=3)
        var.value = row['annot_utt']

    # Adding the specified language to the sheet
    for index, row in df_lang.iterrows():
        # Column 2
        var = worksheet.cell(row=index + 2, column=4)
        var.value = row['utt']
        # Column 3
        var = worksheet.cell(row=index + 2, column=5)
        var.value = row['annot_utt']

    workbook.save('outputs/en-' + language_abbv + '.xlsx')
